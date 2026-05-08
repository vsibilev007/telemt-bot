"""
Экспорт пользователей в CSV и Excel
"""

from __future__ import annotations
import csv
import io
from datetime import UTC, datetime
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def _fmt_bytes(b: int) -> str:
    if not b:
        return "0"
    for unit, thr in [("GB", 1 << 30), ("MB", 1 << 20), ("KB", 1 << 10)]:
        if b >= thr:
            return f"{b / thr:.2f} {unit}"
    return f"{b} B"


def users_to_csv(users: list, server_name: str) -> bytes:
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(
        [
            "username",
            "current_connections",
            "total_octets",
            "total_octets_human",
            "active_unique_ips",
            "recent_unique_ips",
            "max_tcp_conns",
            "max_unique_ips",
            "data_quota_bytes",
            "expiration_rfc3339",
            "server",
        ]
    )
    for u in users:
        writer.writerow(
            [
                u.get("username", ""),
                u.get("current_connections", 0),
                u.get("total_octets", 0),
                _fmt_bytes(u.get("total_octets", 0)),
                u.get("active_unique_ips", 0),
                u.get("recent_unique_ips", 0),
                u.get("max_tcp_conns", ""),
                u.get("max_unique_ips", ""),
                u.get("data_quota_bytes", ""),
                u.get("expiration_rfc3339", ""),
                server_name,
            ]
        )
    return buf.getvalue().encode("utf-8-sig")


def users_to_xlsx(
    users: list, server_name: str, traffic_deltas: list[dict] | None = None
) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Users"
    header_fill = PatternFill("solid", fgColor="1E3A5F")
    header_font = Font(color="FFFFFF", bold=True)
    headers = [
        "Пользователь",
        "Соединений",
        "Трафик всего",
        "Трафик за 7д",
        "Активных IP",
        "Недавних IP",
        "Max TCP",
        "Max IP",
        "Квота",
        "Истекает",
        "Сервер",
    ]
    delta_map: dict[str, int] = {}
    if traffic_deltas:
        delta_map = {r["username"]: r["delta_bytes"] for r in traffic_deltas}
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    green_fill = PatternFill("solid", fgColor="C6EFCE")
    red_fill = PatternFill("solid", fgColor="FFC7CE")
    for row_idx, u in enumerate(users, 2):
        conns = u.get("current_connections", 0)
        exp = u.get("expiration_rfc3339", "")
        row_data = [
            u.get("username", ""),
            conns,
            _fmt_bytes(u.get("total_octets", 0)),
            _fmt_bytes(delta_map.get(u.get("username", ""), 0)),
            u.get("active_unique_ips", 0),
            u.get("recent_unique_ips", 0),
            u.get("max_tcp_conns", ""),
            u.get("max_unique_ips", ""),
            _fmt_bytes(u.get("data_quota_bytes", 0))
            if u.get("data_quota_bytes")
            else "",
            exp[:10] if exp else "",
            server_name,
        ]
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            if conns > 0:
                cell.fill = green_fill
            if exp and exp[:10] < datetime.now(UTC).strftime("%Y-%m-%d"):
                if col_idx == 10:
                    cell.fill = red_fill
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(
            max_len + 2, 40
        )
    ws.freeze_panes = "A2"
    ws2 = wb.create_sheet("Info")
    ws2["A1"] = "Сервер"
    ws2["B1"] = server_name
    ws2["A2"] = "Экспортировано"
    ws2["B2"] = datetime.now(UTC).strftime("%Y-%m-%d %H:%M UTC")
    ws2["A3"] = "Пользователей"
    ws2["B3"] = len(users)
    ws2["A4"] = "Активных"
    ws2["B4"] = sum(1 for u in users if u.get("current_connections", 0) > 0)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
